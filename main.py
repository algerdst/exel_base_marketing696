import datetime
#
import openpyxl
#
import os, glob

path = 'G:\фриланс\Выполненные заказы\garant_exel_parsing\тест файлы'

full_data_dict = {}
data_fields = ['name', 'INN', 'address', 'email', 'phone', 'time_zone', 'director', 'contract_price', 'tender_price',
               'find_by_phrase', 'tender_text', 'customer', 'date_of_contract_public', 'date_of_execute_start',
               'date_of_execute_end'
               ]
doc_count = 1
match=0
count=0
for filename in glob.glob(os.path.join(path, '*.xlsx')):
    book = openpyxl.load_workbook(filename)
    print(f'ОТКРЫЛ ДОКУМЕНТ НОМЕР {doc_count}')
    sheet = book.active
    row = 3
    start = datetime.datetime.now()
    while True:
        # проверка строки на не пустоту
        if not sheet[row][0].value:
            break
        data = {}
        for column in range(15):
            info = sheet[row][column].value
            name = sheet[row][0].value
            try:
                hyperlink = sheet[row][column].hyperlink.target
                data[data_fields[column]] = (sheet[row][column].value, hyperlink)
                continue
            except:
                pass
            data[data_fields[column]] = sheet[row][column].value
        data['email'] = data['email'].split('\n')
        data['phone'] = data['phone'].split('\n')
        count+=1
        print(f'Обработано {count} значений')
        full_data_dict[name] = data
        row += 1
        if len(full_data_dict) % 100 == 0:
            now = datetime.datetime.now()
            print(f"Обработано {len(full_data_dict)}")
            print(f"Собрано 100 значений за {now-start}")
            start = datetime.datetime.now()
    doc_count += 1

print('Найдено всего значений'+str(len(full_data_dict)))
print("Приступаю к записи значений")


fields_names = ['Поставщик / Исполнитель (ссылка)', 'ИНН', 'Адрес поставки', 'Email1', 'Email2', 'Email3',
                'Email4','Email5', 'Телефон1', 'Телефон2', 'Телефон3', 'Телефон4','Телефон5', 'Часовой пояс', 'Руководитель',
                'Цена контракта', 'НМЦК тендера',
                'Найдено по фразе(ссылка)', 'Текст тендера(ссылка)', 'Заказчик (ссылка)', 'Дата публикации контракта',
                'Дата начала исполнения контракта',
                'Дата окончания исполнения контракта'
                ]

book = openpyxl.Workbook()
sheet1 = book.active
for i in range(1,len(fields_names)+1):
    sheet1.cell(1,i).value=fields_names[i-1]


row=2
for key in full_data_dict:
    column=1
    for field in range(15):
        fieldname=data_fields[field]
        if isinstance(full_data_dict[key][fieldname],tuple):
            sheet1.cell(row,column).value=full_data_dict[key][fieldname][0]
            sheet1.cell(row,column).hyperlink=full_data_dict[key][fieldname][1]
            sheet1.cell(row,column).style="Hyperlink"
            column+=1
            continue
        elif fieldname in ['email','phone']:
            for i in range(5):
                try:
                    sheet1.cell(row, column).value = full_data_dict[key][fieldname][i]
                    column+=1
                except:
                    sheet1.cell(row, column).value=''
                    column+=1
        else:
            sheet1.cell(row, column).value = full_data_dict[key][fieldname]
            column+=1
    row+=1

book.save('result_test.xlsx')
book.close()
